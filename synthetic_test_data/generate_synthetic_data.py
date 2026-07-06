#!/usr/bin/env python3
"""Generate randomised, PII-free bank statement regression fixtures.

Design rules
------------
* No word "synthetic/syn/test" anywhere in generated file content.
* Every bank uses its real name, real IFSC prefix, real account-number length,
  and real narration / column conventions.
* Only person-level details (name, account number, branch code, UPI handle)
  are fabricated.
* 150-400 transactions per account over ~12-18 months.
* Each account is assigned ONE format file (not split across four formats).
* Files are named like real bank statements (account-no + date range, etc.).
* synthetic_account_id (SYN0001) appears ONLY in ground_truth.json.
"""
from __future__ import annotations

import argparse
import csv
import json
import random
import shutil
import string
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)


# ---------------------------------------------------------------------------
# Name / location pools (all fabricated but plausible Indian names)
# ---------------------------------------------------------------------------
GIVEN_M = [
    "Rahul", "Amit", "Vijay", "Sanjay", "Suresh", "Ramesh", "Mahesh",
    "Rajesh", "Arun", "Ravi", "Sunil", "Anil", "Vikram", "Arjun", "Deepak",
    "Nikhil", "Rohit", "Prashant", "Vikas", "Gaurav", "Manish", "Rakesh",
    "Pankaj", "Santosh", "Dinesh", "Harish", "Naresh", "Varun", "Tarun",
    "Kapil", "Sachin", "Mohit", "Sumit", "Ankit", "Shivam", "Abhishek",
    "Krishna", "Vishal", "Naveen", "Rajan", "Mohan", "Gopal", "Ashish",
    "Saurabh", "Vivek", "Akash", "Kunal", "Himanshu", "Siddharth", "Kamal",
    "Brijesh", "Yogesh", "Mukesh", "Girish", "Lalit", "Ajay", "Manoj",
]
GIVEN_F = [
    "Priya", "Sunita", "Rekha", "Geeta", "Kavita", "Anita", "Sushma",
    "Meena", "Renu", "Usha", "Asha", "Nisha", "Pooja", "Deepa", "Preeti",
    "Babita", "Sonia", "Ritu", "Swati", "Shweta", "Isha", "Sneha",
    "Pallavi", "Vandana", "Radha", "Kavya", "Divya", "Maya", "Jyoti",
    "Sarita", "Lata", "Mina", "Rani", "Sapna", "Neha", "Shreya", "Poonam",
    "Anjali", "Seema", "Mamta", "Sangita", "Kiran", "Varsha", "Archana",
]
SURNAMES = [
    "Sharma", "Verma", "Singh", "Kumar", "Gupta", "Patel", "Shah", "Mehta",
    "Joshi", "Rao", "Reddy", "Nair", "Pillai", "Iyer", "Agarwal", "Mishra",
    "Tiwari", "Pandey", "Yadav", "Chaudhary", "Saxena", "Srivastava",
    "Dubey", "Shukla", "Bose", "Ghosh", "Das", "Banerjee", "Mukherjee",
    "Chatterjee", "Kapoor", "Malhotra", "Khanna", "Chopra", "Bajaj",
    "Desai", "Jain", "Kulkarni", "Patil", "Naik", "Menon", "Nambiar",
    "Pillai", "Thapar", "Sethi", "Ahuja", "Dutta", "Roy", "Sen",
]

LOCALITIES = [
    "MG Road", "Civil Lines", "Main Market", "Station Road", "Gandhi Nagar",
    "Nehru Place", "Connaught Place", "Karol Bagh", "Lajpat Nagar",
    "Rohini Sector 10", "Pitampura", "Dwarka Sector 6", "Saket",
    "Andheri East", "Bandra West", "Goregaon East", "Malad West",
    "Borivali East", "Koregaon Park", "Baner", "Aundh", "Hinjewadi",
    "Kothrud", "Viman Nagar", "Sector 18 Noida", "Indiranagar",
    "Koramangala", "Whitefield", "Jayanagar", "Basavangudi",
    "Abids", "Banjara Hills", "Himayatnagar", "Ameerpet",
    "Anna Nagar", "T Nagar", "Velachery", "Adyar",
    "Salt Lake Sector 5", "Ballygunge", "Park Street",
    "Hazratganj", "Gomti Nagar", "Alambagh",
    "Vastrapur", "Satellite", "Navrangpura", "Paldi",
    "Mansarovar", "Vaishali Nagar", "Raja Park",
    "Civil Hospital Road", "Sector 43 B", "Sector 17",
]
CITIES = [
    "Mumbai", "Delhi", "Kolkata", "Chennai", "Bangalore", "Hyderabad",
    "Ahmedabad", "Pune", "Jaipur", "Lucknow", "Kanpur", "Nagpur",
    "Indore", "Bhopal", "Patna", "Vadodara", "Surat", "Coimbatore",
    "Kochi", "Chandigarh", "Amritsar", "Ludhiana", "Agra", "Varanasi",
    "Rajkot", "Nashik", "Jodhpur", "Noida", "Gurgaon", "Faridabad",
]

# Merchants / payees for normal transactions
MERCHANTS_UPI = [
    ("Swiggy", "swiggy@icici"), ("Zomato", "zomato@axisbank"),
    ("BigBasket", "bigbasket@ybl"), ("Blinkit", "blinkit@kotak"),
    ("Dunzo", "dunzo@razorpay"), ("Zepto", "zepto@apl"),
    ("Ola Cabs", "olamoney@ola"), ("Uber India", "uber@paytm"),
    ("Amazon Pay", "amazon@apl"), ("Flipkart", "flipkart@ybl"),
    ("PhonePe Merchant", "phonepe@ybl"), ("Paytm Mall", "paytm@paytm"),
    ("BSNL Broadband", "bsnlbroadband@upi"),
    ("Jio Recharge", "jio@paytm"), ("Airtel Payments", "airtel@axis"),
    ("Tata Power", "tatapower@icici"), ("MSEDCL", "msedcl@upi"),
    ("CESC Limited", "cescpayment@upi"), ("BESCOM", "bescompayment@upi"),
    ("Indane Gas", "indanegas@upi"), ("HP Gas", "hpgas@paytm"),
    ("BookMyShow", "bookmyshow@paytm"), ("PVR Cinemas", "pvr@hdfc"),
    ("Netflix India", "netflixindia@icici"),
    ("Hotstar", "hotstar@hdfcbank"), ("Spotify", "spotify@icici"),
    ("MakeMyTrip", "makemytrip@axisbank"), ("IRCTC", "irctc@upi"),
    ("Myntra", "myntra@ybl"), ("Nykaa", "nykaa@razorpay"),
    ("Practo", "practo@paytm"), ("Apollo Pharmacy", "apollopharmacy@ybl"),
]
EMPLOYERS = [
    "Infosys BPO Pvt Ltd", "Tata Consultancy Services",
    "HCL Technologies Ltd", "Wipro Limited", "Tech Mahindra Ltd",
    "Cognizant Technology", "Accenture India Pvt Ltd",
    "Mphasis Limited", "L&T Infotech Ltd", "Hexaware Technologies",
    "Capgemini India", "IBM India Pvt Ltd", "Oracle Financial Services",
    "NIIT Technologies", "Patni Computer Systems",
    "Reliance Retail Ltd", "Vodafone Idea Ltd", "Bharti Airtel Ltd",
    "Maruti Suzuki India", "Bajaj Auto Ltd", "Hero MotoCorp Ltd",
    "HDFC Life Insurance", "ICICI Lombard GIC", "Bajaj Finserv Ltd",
    "Max Healthcare Institute", "Fortis Healthcare Ltd",
    "Dr Reddys Laboratories", "Sun Pharmaceutical Industries",
    "Cipla Limited", "Lupin Limited",
    "Mahindra Logistics Ltd", "Blue Dart Express Ltd",
    "Arvind Ltd", "Raymond Ltd", "Madura Garments Pvt Ltd",
]
INSURANCE_COS = [
    "LIC of India", "HDFC Life Insurance", "ICICI Prudential Life",
    "SBI Life Insurance", "Max Life Insurance", "Bajaj Allianz Life",
    "Reliance Nippon Life", "Tata AIA Life", "Aditya Birla Sun Life",
]
LOAN_INSTITUTIONS = [
    "HDFC Ltd", "LIC Housing Finance", "SBI Home Loans",
    "ICICI Bank Home Loan", "Axis Bank Auto Loan",
    "Bajaj Finance Ltd", "Muthoot Finance", "Manappuram Finance",
    "IDFC First Bank", "Tata Capital Finance",
]

UPI_SUFFIXES = [
    "@oksbi", "@okhdfcbank", "@okaxis", "@okicici", "@ybl",
    "@paytm", "@apl", "@ibl", "@axl", "@upi", "@kotak",
    "@aubank", "@pthdfc", "@ptsbi", "@ptaxis", "@ptyes",
]


# ---------------------------------------------------------------------------
# Bank profiles
# ---------------------------------------------------------------------------
@dataclass
class BankProfile:
    code: str
    name: str
    ifsc_prefix: str  # e.g. "SBIN0"
    acct_len: int
    formats: list[str]          # ordered preference; one is chosen per account
    date_fmt: str               # strftime format used in statements
    pdf_col_style: str          # which PDF layout template to use
    upi_short: str              # short bank code in UPI narrations
    xlsx_col_style: str = ""    # which XLSX column set to use (empty=generic)

    def random_ifsc(self, rng: random.Random) -> str:
        suffix = "".join(rng.choices(string.digits, k=6))
        return self.ifsc_prefix + suffix

    def random_acct(self, rng: random.Random) -> str:
        return "".join(rng.choices(string.digits, k=self.acct_len))

    def random_branch(self, rng: random.Random) -> str:
        return rng.choice(LOCALITIES).upper()

    def fmt_date(self, d: date) -> str:
        return d.strftime(self.date_fmt)


BANKS: dict[str, BankProfile] = {
    "SBI": BankProfile(
        code="SBI", name="STATE BANK OF INDIA",
        ifsc_prefix="SBIN0", acct_len=11,
        formats=["pdf", "xlsx", "csv", "txt"],
        date_fmt="%d-%m-%Y", pdf_col_style="sbi", upi_short="SBIN",
        xlsx_col_style="sbi_xls",
    ),
    "HDFC": BankProfile(
        code="HDFC", name="HDFC BANK LTD",
        ifsc_prefix="HDFC0", acct_len=14,
        formats=["pdf", "xlsx", "csv"],
        date_fmt="%d/%m/%y", pdf_col_style="hdfc", upi_short="HDFC",
    ),
    "AXIS": BankProfile(
        code="AXIS", name="AXIS BANK LIMITED",
        ifsc_prefix="UTIB0", acct_len=15,
        formats=["pdf", "xlsx", "csv"],
        date_fmt="%d-%m-%Y", pdf_col_style="axis", upi_short="UTIB",
    ),
    "KOTAK": BankProfile(
        code="KOTAK", name="KOTAK MAHINDRA BANK LTD",
        ifsc_prefix="KKBK0", acct_len=10,
        formats=["pdf", "csv", "xlsx"],
        date_fmt="%d-%m-%Y", pdf_col_style="kotak", upi_short="KKBK",
    ),
    "PNB": BankProfile(
        code="PNB", name="PUNJAB NATIONAL BANK",
        ifsc_prefix="PUNB0", acct_len=16,
        formats=["pdf", "xlsx", "txt"],
        date_fmt="%d-%m-%Y", pdf_col_style="pnb", upi_short="PUNB",
    ),
    "BANDHAN": BankProfile(
        code="BANDHAN", name="BANDHAN BANK LIMITED",
        ifsc_prefix="BDBL0", acct_len=14,
        formats=["pdf", "csv", "xlsx"],
        date_fmt="%d-%b-%Y", pdf_col_style="bandhan", upi_short="BDBL",
    ),
    "FEDERAL": BankProfile(
        code="FEDERAL", name="THE FEDERAL BANK LIMITED",
        ifsc_prefix="FDRL0", acct_len=14,
        formats=["xlsx", "pdf", "csv"],
        date_fmt="%d-%m-%Y", pdf_col_style="federal", upi_short="FDRL",
        xlsx_col_style="federal_xls",
    ),
    "BOI": BankProfile(
        code="BOI", name="BANK OF INDIA",
        ifsc_prefix="BKID0", acct_len=16,
        formats=["pdf", "xlsx", "csv"],
        date_fmt="%d-%m-%Y", pdf_col_style="boi", upi_short="BKID",
    ),
    "BOB": BankProfile(
        code="BOB", name="BANK OF BARODA",
        ifsc_prefix="BARB0", acct_len=16,
        formats=["pdf", "xlsx", "csv"],
        date_fmt="%d/%m/%Y", pdf_col_style="bob", upi_short="BARB",
    ),
    "UCO": BankProfile(
        code="UCO", name="UCO BANK",
        ifsc_prefix="UCBA0", acct_len=16,
        formats=["pdf", "csv", "xlsx"],
        date_fmt="%d-%m-%Y", pdf_col_style="uco", upi_short="UCBA",
    ),
}

BANK_KEYS = list(BANKS.keys())


# ---------------------------------------------------------------------------
# Pattern metadata (unchanged from original)
# ---------------------------------------------------------------------------
PATTERN_NAMES = {
    1: "duplicate_verification", 2: "failed_reversed_transaction",
    3: "pass_through_routing", 4: "fund_pooling",
    5: "structuring_smurfing", 7: "circular_flow", 8: "money_trail",
    9: "credit_to_cash_out", 10: "cross_statement_links",
    11: "balance_parking", 12: "hub_ranking", 13: "low_value_testing",
    14: "reversal_clusters", 15: "round_value_debit", 16: "shared_upi",
    17: "round_trip", 18: "dormant_reactivation",
    19: "first_contact_large_transfer",
    22: "llm_lead_unknown_shape", 23: "ml_ensemble_unknown_shape",
}
FOLDER_NAMES = {
    1: "pattern_01_duplicate_verification",
    2: "pattern_02_failed_reversed_transaction",
    3: "pattern_03_pass_through_routing",
    4: "pattern_04_fund_pooling",
    5: "pattern_05_structuring_smurfing",
    7: "pattern_07_circular_flow",
    8: "pattern_08_money_trail",
    9: "pattern_09_credit_to_cash_out",
    10: "pattern_10_cross_statement_links",
    11: "pattern_11_balance_parking",
    12: "pattern_12_hub_ranking",
    13: "pattern_13_low_value_testing",
    14: "pattern_14_reversal_clusters",
    15: "pattern_15_round_value_debit",
    16: "pattern_16_shared_upi",
    17: "pattern_17_round_trip",
    18: "pattern_18_dormant_reactivation",
    19: "pattern_19_first_contact_large_transfer",
    22: "pattern_22_llm_lead_unknown_shape",
    23: "pattern_23_ml_ensemble_unknown_shape",
}


def money(v: float | int | Decimal) -> Decimal:
    return Decimal(str(v)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------
@dataclass
class Transaction:
    when: datetime
    direction: str          # "debit" | "credit"
    amount: Decimal
    narration: str
    bank_ref: str
    counterparty: str | None = None
    pattern_ids: set[int] = field(default_factory=set)
    synthetic_ref: str = ""
    balance: Decimal = Decimal("0")


@dataclass
class Account:
    sid: str                # SYN0001 — only in ground_truth.json
    account_no: str         # realistic, per-bank length
    name: str               # realistic Indian name
    ifsc: str
    branch: str
    city: str
    bank: BankProfile
    fmt: str                # the ONE assigned format
    upi_handle: str
    role: str = "subject"
    opening_balance: Decimal = Decimal("0")
    transactions: list[Transaction] = field(default_factory=list)
    files: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Generator
# ---------------------------------------------------------------------------
class Generator:
    def __init__(self, seed: int):
        self.seed = seed
        self.rng = random.Random(seed)
        self._acct_ctr = 0
        self.today = date.today()
        end_offset = self.rng.randint(20, 75)
        self.period_end = self.today - timedelta(days=end_offset)
        self.period_start = self.period_end - timedelta(
            days=self.rng.randint(500, 680)
        )

    # ------------------------------------------------------------------
    # Low-level helpers
    # ------------------------------------------------------------------
    def digits(self, n: int) -> str:
        return "".join(self.rng.choices(string.digits, k=n))

    def alpha(self, n: int) -> str:
        return "".join(self.rng.choices(string.ascii_uppercase, k=n))

    def ref(self, prefix: str = "", length: int = 12) -> str:
        pad = max(1, length - len(prefix))
        return prefix + self.digits(pad)

    def dt(self, d: date, odd: bool = False) -> datetime:
        h = self.rng.choice((1, 2, 3, 22, 23)) if odd else self.rng.randint(7, 21)
        return datetime.combine(d, time(h, self.rng.randint(0, 59), self.rng.randint(0, 59)))

    def amount(self, lo: int, hi: int, step: int = 1, paise: bool = True) -> Decimal:
        v = self.rng.randrange(lo, hi + 1, step)
        if paise:
            v += self.rng.randint(0, 99) / 100
        return money(v)

    def _pick_name(self) -> tuple[str, str]:
        if self.rng.random() < 0.55:
            return self.rng.choice(GIVEN_M), self.rng.choice(SURNAMES)
        return self.rng.choice(GIVEN_F), self.rng.choice(SURNAMES)

    def _upi_handle(self, first: str, last: str, bank: BankProfile) -> str:
        style = self.rng.randint(0, 3)
        if style == 0:
            base = f"{first.lower()}.{last.lower()}"
        elif style == 1:
            base = f"{first.lower()}{last.lower()[:4]}{self.rng.randint(10,99)}"
        elif style == 2:
            base = self.digits(10)
        else:
            base = f"{first.lower()}{self.rng.randint(100,999)}"
        suffix = self.rng.choice(UPI_SUFFIXES)
        return f"{base}{suffix}"

    def _narr_upi(self, direction: str, ref12: str, counterparty_name: str,
                  cp_bank_short: str, cp_handle: str, bank: BankProfile) -> str:
        style = bank.pdf_col_style
        cname = counterparty_name.split()[0] if counterparty_name else "Payee"
        if style == "hdfc":
            dr_cr = "DR" if direction == "debit" else "CR"
            return (f"UPI-{counterparty_name}-{cp_handle}-"
                    f"{ref12}-{dr_cr} VIA UPI")
        if style == "axis":
            tag = "P2A" if direction == "credit" else "P2M"
            return f"UPI{tag}{ref12}{counterparty_name.replace(' ','')[:12]}UPI{cp_bank_short}"
        if style in ("sbi",):
            dr_cr = "DR" if direction == "debit" else "CR"
            return (f"UPI/{dr_cr}/{ref12}/{cname}/{cp_bank_short}/"
                    f"{cp_handle}/UPI")
        # kotak / federal / pnb / boi / bob / uco — use sbi-like format
        dr_cr = "DR" if direction == "debit" else "CR"
        return (f"UPI/{dr_cr}/{ref12}/{cname}/{cp_bank_short}/"
                f"{cp_handle}/UPI")

    def _narr_neft(self, direction: str, ref: str, party_name: str,
                   party_ifsc: str, bank: BankProfile) -> str:
        style = bank.pdf_col_style
        if style == "hdfc":
            if direction == "credit":
                return f"NEFT*{ref}/{party_name[:20]}/CR"
            return f"NEFT*{ref}/{party_name[:20]}/DR"
        if style == "sbi":
            if direction == "credit":
                return f"INB NEFT-{ref}/{party_name[:20]}/{party_ifsc[:4]}/CR"
            return f"INB NEFT-{ref}/{party_name[:20]}/{party_ifsc[:4]}/DR"
        if style == "axis":
            return f"NEFT/{ref}/{party_name[:20]}/{party_ifsc}"
        return f"NEFT/{ref}/{party_name[:20]}"

    def _narr_imps(self, direction: str, ref12: str, party_name: str,
                   party_bank: str, bank: BankProfile) -> str:
        style = bank.pdf_col_style
        if style == "hdfc":
            return (f"IMPS-{ref12}-{party_name[:16]}-"
                    f"{party_bank[:4]}-XXXXXXX{self.digits(4)}-IMPSTXN")
        if style == "sbi":
            return f"INB IMPS/{ref12}/{party_name[:16]}/{party_bank[:4]}"
        if style == "federal":
            tag = "P2A" if direction == "credit" else "P2M"
            return f"IMPS/{tag}/{ref12}//{party_name[:16]}/{party_bank[:4]}"
        if style == "axis":
            return f"IMPS/{ref12}/{party_name[:16]}/{party_bank[:4]}"
        return f"IMPS-{ref12}-{party_name[:16]}"

    def _narr_rtgs(self, direction: str, ref: str, party_name: str,
                   party_ifsc: str, bank: BankProfile) -> str:
        style = bank.pdf_col_style
        if style == "hdfc":
            if direction == "credit":
                return (f"RTGS CR-{party_ifsc[:8]}-{party_name[:16]}-{ref}")
            return f"RTGS-{ref}-{party_name[:16]}/{party_ifsc}"
        if style == "sbi":
            return f"SBIRTGS{ref}/{party_ifsc}/{party_name[:16]}"
        return f"RTGS/{ref}/{party_name[:16]}/{party_ifsc}"

    def _narr_atm(self, acct: Account) -> str:
        style = acct.bank.pdf_col_style
        city = acct.city.upper()
        loc = acct.branch.replace(" ", "")[:8].upper()
        seq = self.digits(12)
        if style == "hdfc":
            card = f"XXXXXX{self.digits(4)}"
            loc_code = self.alpha(3) + self.digits(5)
            return f"ATW-{card}-{loc_code}-{city} {seq}"
        if style == "sbi":
            branch_code = self.digits(5)
            return f"ATM WDL\nATM CASH {branch_code} {acct.branch[:20].upper()} {city}"
        if style == "axis":
            prefix = "ATMCASHAXIS" + self.alpha(3)
            date_str = self.rng.choice(["", ""])  # axis has no date in ATM narr
            return f"{prefix}{seq[:8]}{city[:6].upper()}"
        if style in ("kotak",):
            machine = self.alpha(3) + self.digits(6)
            return f"ATM-CASH-WDL/{machine}/{city}"
        branch_code = self.digits(5)
        return f"ATM CASH WITHDRAWAL/{branch_code}/{city}"

    def _narr_salary(self, employer: str, ref: str, bank: BankProfile) -> str:
        style = bank.pdf_col_style
        if style == "hdfc":
            return f"SALARY-{ref}/NEFT/{employer[:24]}"
        if style == "sbi":
            return f"INB NEFT-{ref}/{employer[:20]}/SALARY CR"
        if style == "axis":
            return f"SALARY/{ref}/{employer[:24]}"
        return f"SALARY CREDIT/{ref}/{employer[:24]}"

    def _narr_emi(self, institution: str, loan_ref: str, bank: BankProfile) -> str:
        if bank.pdf_col_style == "hdfc":
            return f"ECS-NACH-{loan_ref}-{institution[:20]}"
        return f"NACH/MANDATE/{loan_ref}/{institution[:20]}"

    # ------------------------------------------------------------------
    # Account creation
    # ------------------------------------------------------------------
    def new_account(self, role: str = "subject", normal: bool = True,
                    bank_code: str | None = None) -> Account:
        self._acct_ctr += 1
        sid = f"SYN{self._acct_ctr:04d}"
        bp = BANKS[bank_code] if bank_code else BANKS[self.rng.choice(BANK_KEYS)]
        first, last = self._pick_name()
        full_name = f"{first} {last}".upper()
        city = self.rng.choice(CITIES)
        upi = self._upi_handle(first, last, bp)
        fmt = self.rng.choice(bp.formats)
        acct = Account(
            sid=sid,
            account_no=bp.random_acct(self.rng),
            name=full_name,
            ifsc=bp.random_ifsc(self.rng),
            branch=bp.random_branch(self.rng),
            city=city,
            bank=bp,
            fmt=fmt,
            upi_handle=upi,
            role=role,
            opening_balance=self.amount(20_000, 180_000, 500),
        )
        if normal:
            self.add_normal_activity(acct)
        return acct

    # ------------------------------------------------------------------
    # Normal activity: 150-400 transactions
    # ------------------------------------------------------------------
    def add_normal_activity(self, acct: Account, sparse: bool = False) -> None:
        rng = self.rng
        bp = acct.bank
        employer = rng.choice(EMPLOYERS)
        salary = self.amount(38_000, 145_000, 500)
        ins_co = rng.choice(INSURANCE_COS)
        ins_ref = self.ref("LIC", 14)
        ins_premium = self.amount(1_800, 8_500, 100, False)
        ins_freq = rng.choice([1, 3])  # monthly or quarterly
        loan_inst = rng.choice(LOAN_INSTITUTIONS) if rng.random() < 0.6 else None
        loan_ref = self.ref("LN", 14)
        emi_amount = self.amount(4_500, 22_000, 500, False) if loan_inst else Decimal(0)
        sip_inst = "Mirae Asset MF" if rng.random() < 0.5 else "Axis Mutual Fund"
        sip_amount = self.amount(2_000, 15_000, 500, False)
        do_sip = rng.random() < 0.55
        rent_amount = self.amount(6_000, 28_000, 500, False)
        mobile_op = rng.choice(["Jio Recharge", "Airtel Payments", "Vi Mobile"])
        mobile_handle = rng.choice(["jio@paytm", "airtel@axis", "vi@airtelaxis"])
        broadband = rng.choice(["ACT Fibernet", "Hathway", "BSNL Broadband"])
        elec_board = rng.choice(["MSEDCL", "BESCOM", "TSSPDCL", "CESC Limited", "KSEB"])
        elec_handle = f"{elec_board.split()[0].lower()}payment@upi"

        # Determine salary day (1-7 each month, with jitter)
        sal_day_base = rng.randint(1, 7)

        def _sal_day(month_date: date) -> date:
            d = sal_day_base + rng.randint(-1, 2)
            d = max(1, min(d, 7))
            return date(month_date.year, month_date.month, d)

        cursor = self.period_start + timedelta(days=rng.randint(1, 14))
        month = date(cursor.year, cursor.month, 1)

        while month <= self.period_end:
            skip = sparse and rng.random() < 0.35
            if not skip:
                # --- Salary credit ---
                sal_date = _sal_day(month)
                if sal_date <= self.period_end:
                    sal_var = rng.randint(-2000, 3500)
                    sal_actual = money(salary + sal_var)
                    if sal_actual > 0:
                        self._add_tx(acct, sal_date, "credit", sal_actual,
                                     self._narr_salary(employer, self.ref("S", 14), bp))

                # --- Rent (NACH, 1-5 of month) ---
                rent_d = sal_date + timedelta(days=rng.randint(1, 5))
                if rent_d <= self.period_end:
                    self._add_tx(acct, rent_d, "debit", rent_amount,
                                 self._narr_emi(f"RENT MANDATE", self.ref("RNT", 10), bp))

                # --- EMI ---
                if loan_inst:
                    emi_d = sal_date + timedelta(days=rng.randint(2, 8))
                    if emi_d <= self.period_end:
                        self._add_tx(acct, emi_d, "debit", emi_amount,
                                     self._narr_emi(loan_inst, loan_ref, bp))

                # --- Insurance premium ---
                if month.month % ins_freq == 0:
                    ins_d = month + timedelta(days=rng.randint(3, 12))
                    if ins_d <= self.period_end:
                        upi_ref = self.digits(12)
                        narr = self._narr_upi("debit", upi_ref,
                                              ins_co, "LICI",
                                              "licindia@licindia", bp)
                        self._add_tx(acct, ins_d, "debit", ins_premium, narr)

                # --- SIP ---
                if do_sip:
                    sip_d = month + timedelta(days=rng.randint(10, 18))
                    if sip_d <= self.period_end:
                        self._add_tx(acct, sip_d, "debit", sip_amount,
                                     self._narr_emi(sip_inst, self.ref("SIP", 10), bp))

                # --- Utility: electricity ---
                elec_d = month + timedelta(days=rng.randint(8, 22))
                if elec_d <= self.period_end:
                    self._add_tx(acct, elec_d, "debit",
                                 self.amount(600, 4_200, 50),
                                 self._narr_upi("debit", self.digits(12),
                                                elec_board, "UPIN",
                                                elec_handle, bp))

                # --- Gas bill (monthly) ---
                gas_d = month + timedelta(days=rng.randint(10, 25))
                if gas_d <= self.period_end:
                    self._add_tx(acct, gas_d, "debit",
                                 self.amount(800, 2_800, 50),
                                 self._narr_upi("debit", self.digits(12),
                                                "Indane Gas", "ICICI",
                                                "indanegas@upi", bp))

                # --- Mobile recharge (2-5 per month) ---
                for _ in range(rng.randint(2, 5)):
                    mob_d = month + timedelta(days=rng.randint(1, 28))
                    if mob_d <= self.period_end:
                        self._add_tx(acct, mob_d, "debit",
                                     self.amount(179, 999, 10, False),
                                     self._narr_upi("debit", self.digits(12),
                                                    mobile_op, "PAYTM",
                                                    mobile_handle, bp))

                # --- Broadband / DTH ---
                if rng.random() < 0.75:
                    bb_d = month + timedelta(days=rng.randint(1, 28))
                    if bb_d <= self.period_end:
                        self._add_tx(acct, bb_d, "debit",
                                     self.amount(299, 1_299, 50, False),
                                     self._narr_upi("debit", self.digits(12),
                                                    broadband, "RAZORPAY",
                                                    "broadband@paytm", bp))

                # --- Food delivery & grocery (8-18 per month) ---
                for _ in range(rng.randint(8, 18)):
                    fd_d = month + timedelta(days=rng.randint(0, 30))
                    if fd_d <= self.period_end:
                        merchant, mhandle = rng.choice(MERCHANTS_UPI[:14])
                        self._add_tx(acct, fd_d, "debit",
                                     self.amount(120, 2_800, 10),
                                     self._narr_upi("debit", self.digits(12),
                                                    merchant, "ICICI",
                                                    mhandle, bp))

                # --- ATM withdrawal (2-6 per month) ---
                for _ in range(rng.randint(2, 6)):
                    atm_d = month + timedelta(days=rng.randint(0, 30))
                    if atm_d <= self.period_end:
                        self._add_tx(acct, atm_d, "debit",
                                     self.amount(1_000, 10_000, 500, False),
                                     self._narr_atm(acct))

                # --- POS / shopping (3-8 per month) ---
                for _ in range(rng.randint(3, 8)):
                    pos_d = month + timedelta(days=rng.randint(0, 30))
                    if pos_d <= self.period_end:
                        mname = rng.choice([
                            "RELIANCE MART", "D MART", "MORE SUPER MARKET",
                            "VISHAL MEGA MART", "LIFESTYLE STORES",
                            "CENTRAL MALL", "V MART", "SHOPPERS STOP",
                            "BIG BAZAAR", "STAR BAZAAR", "METRO CASH CARRY",
                            "BRAND FACTORY", "MAX FASHION", "PANTALOONS",
                        ])
                        self._add_tx(acct, pos_d, "debit",
                                     self.amount(300, 8_500, 50),
                                     f"POS {mname} {acct.city.upper()[:10]}")

                # --- Fuel / petrol (2-4 per month) ---
                for _ in range(rng.randint(2, 4)):
                    fuel_d = month + timedelta(days=rng.randint(0, 30))
                    if fuel_d <= self.period_end:
                        pump = rng.choice([
                            "HP PETROL PUMP", "INDIAN OIL", "BHARAT PETROLEUM",
                            "SHELL FUEL STATION", "RELIANCE PETROLEUM",
                        ])
                        self._add_tx(acct, fuel_d, "debit",
                                     self.amount(800, 4_500, 50),
                                     f"POS {pump} {acct.city.upper()[:8]}")

                # --- P2P UPI transfers (10-22 per month) ---
                for _ in range(rng.randint(10, 22)):
                    p2p_d = month + timedelta(days=rng.randint(0, 30))
                    if p2p_d <= self.period_end:
                        direction = "debit" if rng.random() < 0.55 else "credit"
                        cp_first, cp_last = self._pick_name()
                        cp_name = f"{cp_first} {cp_last}"
                        cp_bank = rng.choice(BANK_KEYS)
                        cp_handle = self._upi_handle(cp_first, cp_last, BANKS[cp_bank])
                        self._add_tx(acct, p2p_d, direction,
                                     self.amount(100, 22_000, 50),
                                     self._narr_upi(direction, self.digits(12),
                                                    cp_name, BANKS[cp_bank].upi_short,
                                                    cp_handle, bp))

                # --- NEFT inward (freelance / bonus) ---
                if rng.random() < 0.4:
                    neft_d = month + timedelta(days=rng.randint(5, 25))
                    if neft_d <= self.period_end:
                        sender = rng.choice(EMPLOYERS)
                        self._add_tx(acct, neft_d, "credit",
                                     self.amount(5_000, 55_000, 500),
                                     self._narr_neft("credit",
                                                     self.ref("N", 14), sender,
                                                     rng.choice(list(BANKS.values())).random_ifsc(rng),
                                                     bp))

                # --- IMPS outward (1-2 per month) ---
                for _ in range(rng.randint(1, 2)):
                    if rng.random() < 0.5:
                        imps_d = month + timedelta(days=rng.randint(0, 28))
                        if imps_d <= self.period_end:
                            cp_first, cp_last = self._pick_name()
                            cp_bank = rng.choice(BANK_KEYS)
                            self._add_tx(acct, imps_d, "debit",
                                         self.amount(500, 45_000, 100),
                                         self._narr_imps("debit", self.digits(12),
                                                         f"{cp_first} {cp_last}",
                                                         BANKS[cp_bank].upi_short, bp))

                # --- Medical / pharmacy ---
                if rng.random() < 0.5:
                    med_d = month + timedelta(days=rng.randint(0, 28))
                    if med_d <= self.period_end:
                        med_m, med_h = rng.choice([
                            ("Apollo Pharmacy", "apollopharmacy@ybl"),
                            ("MedPlus Health", "medplus@icici"),
                            ("Netmeds", "netmeds@paytm"),
                            ("1mg", "tatahealth@icici"),
                        ])
                        self._add_tx(acct, med_d, "debit",
                                     self.amount(150, 3_500, 10),
                                     self._narr_upi("debit", self.digits(12),
                                                    med_m, "ICICI", med_h, bp))

                # --- Online shopping (1-3 per month) ---
                for _ in range(rng.randint(1, 3)):
                    shop_d = month + timedelta(days=rng.randint(0, 28))
                    if shop_d <= self.period_end:
                        shop_m, shop_h = rng.choice([
                            ("Amazon Pay", "amazon@apl"),
                            ("Flipkart", "flipkart@ybl"),
                            ("Myntra", "myntra@ybl"),
                            ("Meesho", "meesho@ybl"),
                            ("Nykaa", "nykaa@razorpay"),
                        ])
                        self._add_tx(acct, shop_d, "debit",
                                     self.amount(300, 6_000, 50),
                                     self._narr_upi("debit", self.digits(12),
                                                    shop_m, "PAYTM", shop_h, bp))

                # --- Bank charges / GST ---
                if rng.random() < 0.5:
                    chrg_d = month + timedelta(days=rng.randint(25, 30))
                    if chrg_d <= self.period_end:
                        chrg = self.amount(50, 590, 10, False)
                        gst = money(chrg * Decimal("0.18"))
                        narr_map = {
                            "sbi": "CHG FOR NEFT/IMPS TRANSACTION",
                            "hdfc": "MOBILE BANKING CHARGES",
                            "axis": "Monthly Service Chrgs",
                            "kotak": "ACCOUNT MAINTENANCE CHARGES",
                            "pnb": "SMS ALERT CHARGES",
                            "bandhan": "ACCOUNT MAINTENANCE FEE",
                            "federal": "TRANSACTION CHARGES",
                            "boi": "SERVICE CHARGES",
                            "bob": "SMS CHARGES",
                            "uco": "MAINTENANCE CHARGES",
                        }
                        self._add_tx(acct, chrg_d, "debit", chrg,
                                     narr_map.get(bp.pdf_col_style, "BANK CHARGES"))
                        if gst > 0:
                            self._add_tx(acct, chrg_d, "debit", gst,
                                         "GST ON CHARGES")

                # --- Quarterly interest credit ---
                if month.month in (3, 6, 9, 12):
                    int_d = date(month.year, month.month,
                                 min(28, rng.randint(25, 31)))
                    if int_d <= self.period_end:
                        self._add_tx(acct, int_d, "credit",
                                     self.amount(10, 1_200, 5),
                                     "INTEREST CREDIT")

                # --- Entertainment subscriptions (1-2 per month) ---
                for _ in range(rng.randint(1, 2)):
                    if rng.random() < 0.6:
                        sub_d = month + timedelta(days=rng.randint(1, 28))
                        if sub_d <= self.period_end:
                            svc, shandle = rng.choice(MERCHANTS_UPI[24:32])
                            self._add_tx(acct, sub_d, "debit",
                                         self.amount(149, 999, 50, False),
                                         self._narr_upi("debit", self.digits(12),
                                                        svc, "ICICI", shandle, bp))

                # --- Ride hailing (3-6 per month) ---
                for _ in range(rng.randint(3, 6)):
                    ride_d = month + timedelta(days=rng.randint(0, 28))
                    if ride_d <= self.period_end:
                        ride_m, ride_h = rng.choice([
                            ("Ola Cabs", "olamoney@ola"),
                            ("Uber India", "uber@paytm"),
                            ("Rapido", "rapido@icici"),
                        ])
                        self._add_tx(acct, ride_d, "debit",
                                     self.amount(80, 900, 10),
                                     self._narr_upi("debit", self.digits(12),
                                                    ride_m, "PAYTM", ride_h, bp))

            # Advance to next month
            if month.month == 12:
                month = date(month.year + 1, 1, 1)
            else:
                month = date(month.year, month.month + 1, 1)

    def _add_tx(self, acct: Account, d: date, direction: str, amount: Decimal,
                narration: str, pattern: int | None = None,
                ref: str | None = None,
                counterparty: str | None = None,
                odd_time: bool = False) -> Transaction:
        tx = Transaction(
            when=self.dt(d, odd_time),
            direction=direction,
            amount=money(amount),
            narration=narration,
            bank_ref=ref or self.ref("", 12),
            counterparty=counterparty,
        )
        if pattern is not None:
            tx.pattern_ids.add(pattern)
        acct.transactions.append(tx)
        return tx

    def tx(self, acct: Account, d: date, direction: str, amount: Decimal,
           narration: str, pattern: int | None = None,
           ref: str | None = None,
           counterparty: str | None = None,
           odd_time: bool = False) -> Transaction:
        return self._add_tx(acct, d, direction, money(amount), narration,
                            pattern, ref, counterparty, odd_time)

    def transfer(self, src: Account, dst: Account, d: date, amount: Decimal,
                 pattern: int, channel: str | None = None,
                 same_ref: bool = True) -> tuple[Transaction, Transaction]:
        channel = channel or self.rng.choice(("NEFT", "IMPS", "UPI", "RTGS"))
        ref = self.ref(channel[0], 12)
        dst_ref = ref if same_ref else self.ref("X", 12)

        if channel == "UPI":
            out_narr = self._narr_upi("debit", self.digits(12),
                                      dst.name, dst.bank.upi_short,
                                      dst.upi_handle, src.bank)
            in_narr = self._narr_upi("credit", self.digits(12),
                                     src.name, src.bank.upi_short,
                                     src.upi_handle, dst.bank)
        elif channel == "NEFT":
            out_narr = self._narr_neft("debit", ref, dst.name, dst.ifsc, src.bank)
            in_narr = self._narr_neft("credit", dst_ref, src.name, src.ifsc, dst.bank)
        elif channel == "IMPS":
            out_narr = self._narr_imps("debit", ref, dst.name,
                                       dst.bank.upi_short, src.bank)
            in_narr = self._narr_imps("credit", dst_ref, src.name,
                                      src.bank.upi_short, dst.bank)
        else:  # RTGS
            out_narr = self._narr_rtgs("debit", ref, dst.name, dst.ifsc, src.bank)
            in_narr = self._narr_rtgs("credit", dst_ref, src.name, src.ifsc, dst.bank)

        debit = self.tx(src, d, "debit", amount, out_narr, pattern, ref, dst.sid)
        credit = self.tx(dst, d, "credit", amount, in_narr, pattern, dst_ref, src.sid)
        return debit, credit

    # ------------------------------------------------------------------
    # Pattern planting (realistic narrations, no SYNTHETIC strings)
    # ------------------------------------------------------------------
    def event_day(self, lo: int = 80, hi: int = 300) -> date:
        span = (self.period_end - self.period_start).days
        offset = self.rng.randint(min(lo, span // 3), min(hi, span - 30))
        return self.period_start + timedelta(days=offset)

    def plant(self, pid: int, accounts: list[Account]) -> dict[str, Any]:  # noqa: C901
        d = self.event_day()
        a = accounts[0]
        txs: list[Transaction] = []
        notes = ""
        tier = "strong"

        if pid == 1:
            amt = self.amount(6_000, 42_000, 100)
            ref = self.ref("I", 12)
            cp_first, cp_last = self._pick_name()
            cp_name = f"{cp_first} {cp_last}"
            narr = self._narr_imps("debit", ref, cp_name,
                                   self.rng.choice(BANK_KEYS), a.bank)
            orig = self.tx(a, d, "debit", amt, narr, pid, ref)
            dup = Transaction(orig.when, orig.direction, orig.amount,
                              orig.narration, orig.bank_ref,
                              orig.counterparty, {pid})
            a.transactions.append(dup)
            txs = [orig, dup]
            notes = "Exact duplicate row embedded in realistic surrounding activity."

        elif pid == 2:
            amt = self.amount(8_000, 75_000, 100)
            ref = self.ref("I", 12)
            cp_first, cp_last = self._pick_name()
            narr_out = self._narr_neft("debit", ref, f"{cp_first} {cp_last}",
                                       self.rng.choice(list(BANKS.values())).random_ifsc(self.rng),
                                       a.bank)
            debit = self.tx(a, d, "debit", amt, narr_out, pid, ref)
            rev_narr = f"RETURN/{ref}/TRANSACTION FAILED"
            reversal = self.tx(a, d + timedelta(days=self.rng.randint(1, 2)),
                               "credit", amt, rev_narr, pid)
            txs = [debit, reversal]
            notes = "Debit followed by credit reversal for the exact same amount."

        elif pid == 3:
            inbound = []
            total = Decimal("0")
            for i in range(self.rng.randint(4, 6)):
                amt = self.amount(35_000, 95_000, 100)
                total += amt
                cp_first, cp_last = self._pick_name()
                channel = self.rng.choice(("NEFT", "IMPS", "UPI"))
                if channel == "UPI":
                    narr = self._narr_upi(
                        "credit", self.digits(12), f"{cp_first} {cp_last}",
                        self.rng.choice(BANK_KEYS),
                        self._upi_handle(cp_first, cp_last, BANKS[self.rng.choice(BANK_KEYS)]),
                        a.bank)
                else:
                    narr = self._narr_neft(
                        "credit", self.ref("N", 12), f"{cp_first} {cp_last}",
                        BANKS[self.rng.choice(BANK_KEYS)].random_ifsc(self.rng),
                        a.bank)
                inbound.append(self.tx(a, d + timedelta(hours=i), "credit", amt, narr, pid))
            dst_first, dst_last = self._pick_name()
            out_amt = money(total * Decimal(str(self.rng.uniform(.88, .96))))
            out = self.tx(a, d + timedelta(days=self.rng.randint(1, 2)), "debit",
                          out_amt,
                          self._narr_rtgs("debit", self.ref("R", 12),
                                          f"{dst_first} {dst_last}",
                                          BANKS[self.rng.choice(BANK_KEYS)].random_ifsc(self.rng),
                                          a.bank),
                          pid)
            txs = inbound + [out]
            notes = "Multiple unrelated inbound credits followed by rapid onward routing."

        elif pid == 4:
            for i in range(self.rng.randint(5, 8)):
                cp_first, cp_last = self._pick_name()
                cp_bank = self.rng.choice(BANK_KEYS)
                narr = self._narr_upi("credit", self.digits(12),
                                      f"{cp_first} {cp_last}",
                                      BANKS[cp_bank].upi_short,
                                      self._upi_handle(cp_first, cp_last, BANKS[cp_bank]),
                                      a.bank)
                txs.append(self.tx(a, d + timedelta(days=self.rng.randint(0, 4), hours=i),
                                   "credit", self.amount(18_000, 72_000, 100), narr, pid))
            notes = "Fund pooling from multiple unrelated senders within a short window."

        elif pid == 5:
            branch_city = a.city.upper()
            for i in range(self.rng.randint(5, 8)):
                narr = (f"CASH DEPOSIT/{self.digits(10)}/"
                        f"{a.branch[:16].upper()}/{branch_city}")
                txs.append(self.tx(a, d + timedelta(days=i), "credit",
                                   self.amount(42_000, 49_900, 100, False),
                                   narr, pid))
            notes = "Repeated cash deposits just below a common reporting threshold."

        elif pid == 7:
            amt = self.amount(145_000, 360_000, 1000)
            for i in range(len(accounts)):
                src, dst = accounts[i], accounts[(i + 1) % len(accounts)]
                leg = money(amt * Decimal(str(self.rng.uniform(.94, .995)))) if i else amt
                ch = self.rng.choice(("NEFT", "IMPS", "RTGS"))
                txs.extend(self.transfer(src, dst, d + timedelta(days=i), leg, pid, ch))
            notes = "Closed-loop circular flow; every hop corroborated by both statements."

        elif pid == 8:
            amt = self.amount(120_000, 410_000, 1000)
            for i in range(len(accounts) - 1):
                amt = money(amt * Decimal(str(self.rng.uniform(.86, .97))))
                ch = self.rng.choice(("NEFT", "IMPS", "RTGS"))
                txs.extend(self.transfer(accounts[i], accounts[i + 1],
                                         d + timedelta(days=i), amt, pid, ch))
            notes = "Multi-hop trail; real entries on both sides of every transfer."

        elif pid == 9:
            amt = self.amount(90_000, 280_000, 1000)
            cp_first, cp_last = self._pick_name()
            credit = self.tx(a, d, "credit", amt,
                             self._narr_neft("credit", self.ref("N", 12),
                                             f"{cp_first} {cp_last}",
                                             BANKS[self.rng.choice(BANK_KEYS)].random_ifsc(self.rng),
                                             a.bank),
                             pid)
            cash_amt = money(amt * Decimal(str(self.rng.uniform(.75, .94))))
            cash = self.tx(a, d + timedelta(days=self.rng.randint(0, 2)),
                           "debit", cash_amt, self._narr_atm(a), pid)
            txs = [credit, cash]
            notes = "Large inward credit followed promptly by near-equivalent ATM withdrawal."

        elif pid == 10:
            ch = self.rng.choice(("NEFT", "IMPS"))
            txs.extend(self.transfer(accounts[0], accounts[1], d,
                                     self.amount(55_000, 240_000, 1000), pid, ch, True))
            notes = "Same bank reference appears on two independent account statements."

        elif pid == 11:
            cp_first, cp_last = self._pick_name()
            txs = [self.tx(a, d, "credit", self.amount(260_000, 720_000, 1000),
                           self._narr_rtgs("credit", self.ref("R", 12),
                                           f"{cp_first} {cp_last}",
                                           BANKS[self.rng.choice(BANK_KEYS)].random_ifsc(self.rng),
                                           a.bank),
                           pid)]
            notes = "Large credit remains parked; subsequent activity is minor."

        elif pid == 12:
            for i, spoke in enumerate(accounts[1:]):
                ch = self.rng.choice(("UPI", "IMPS", "NEFT"))
                txs.extend(self.transfer(spoke, a, d + timedelta(days=i % 4),
                                         self.amount(14_000, 68_000, 100), pid, ch))
            notes = "Hub receives from 6-8 spoke accounts with corroborating statements."

        elif pid == 13:
            for i in range(self.rng.randint(3, 5)):
                txs.extend(self.transfer(accounts[i % 2], accounts[(i + 1) % 2],
                                         d + timedelta(days=i),
                                         self.amount(1, 35, 1), pid, "UPI"))
            notes = "Reciprocal low-value probes on both real account sides."

        elif pid == 14:
            for i in range(self.rng.randint(3, 5)):
                cp_first, cp_last = self._pick_name()
                amt = self.amount(4_000, 36_000, 100)
                ref = self.ref("U", 12)
                debit = self.tx(a, d + timedelta(days=i * 3), "debit", amt,
                                self._narr_upi("debit", self.digits(12),
                                               f"{cp_first} {cp_last}",
                                               self.rng.choice(BANK_KEYS),
                                               f"{cp_first.lower()}@paytm", a.bank),
                                pid, ref)
                txs.extend([debit,
                             self.tx(a, debit.when.date() + timedelta(days=1),
                                     "credit", amt, f"RETURN/{ref}/REVERSAL", pid)])
            notes = "Repeated debit-reversal pattern across multiple cycles."

        elif pid == 15:
            cp_first, cp_last = self._pick_name()
            for i in range(self.rng.randint(4, 7)):
                amt = self.amount(10_000, 90_000, 5_000, False)
                txs.append(self.tx(a, d + timedelta(days=i * 2), "debit", amt,
                                   self._narr_neft("debit", self.ref("N", 12),
                                                   f"{cp_first} {cp_last}",
                                                   BANKS[self.rng.choice(BANK_KEYS)].random_ifsc(self.rng),
                                                   a.bank),
                                   pid))
            tier = "weak"
            notes = "Cluster of round-value outward transfers amid non-round routine spending."

        elif pid == 16:
            shared_handle = (f"{self.rng.choice(GIVEN_M).lower()}"
                             f"{self.rng.randint(10,99)}"
                             f"{self.rng.choice(UPI_SUFFIXES)}")
            for i, acct_i in enumerate(accounts):
                txs.append(self.tx(acct_i, d + timedelta(days=i), "debit",
                                   self.amount(750, 8_500, 10),
                                   self._narr_upi("debit", self.digits(12),
                                                  "Common Payee",
                                                  self.rng.choice(BANK_KEYS),
                                                  shared_handle, acct_i.bank),
                                   pid))
            notes = f"Handle {shared_handle} appears across separate account statements."

        elif pid == 17:
            amt = self.amount(180_000, 420_000, 1000)
            txs.extend(self.transfer(accounts[0], accounts[1], d, amt, pid, "NEFT"))
            returned = money(amt * Decimal(str(self.rng.uniform(.87, .96))))
            if len(accounts) == 3:
                txs.extend(self.transfer(accounts[1], accounts[2],
                                         d + timedelta(days=2), returned, pid, "IMPS"))
                txs.extend(self.transfer(accounts[2], accounts[0],
                                         d + timedelta(days=self.rng.randint(3, 6)),
                                         money(returned * Decimal(str(self.rng.uniform(.96, .99)))),
                                         pid, "RTGS"))
            else:
                txs.extend(self.transfer(accounts[1], accounts[0],
                                         d + timedelta(days=self.rng.randint(2, 6)),
                                         returned, pid, "IMPS"))
            notes = "Out-and-back round-trip via different channels; all sides have statements."

        elif pid == 18:
            a.transactions.clear()
            early = self.period_start + timedelta(days=self.rng.randint(5, 25))
            txs.append(self.tx(a, early, "credit", self.amount(800, 4_500, 10),
                                "INTEREST CREDIT", pid))
            reactivate = early + timedelta(days=self.rng.randint(220, 310))
            cp_first, cp_last = self._pick_name()
            txs.append(self.tx(a, reactivate, "credit",
                                self.amount(95_000, 340_000, 1000),
                                self._narr_neft("credit", self.ref("N", 12),
                                                f"{cp_first} {cp_last}",
                                                BANKS[self.rng.choice(BANK_KEYS)].random_ifsc(self.rng),
                                                a.bank),
                                pid))
            cp2_first, cp2_last = self._pick_name()
            txs.append(self.tx(a, reactivate + timedelta(days=1), "debit",
                                self.amount(55_000, 88_000, 1000),
                                self._narr_imps("debit", self.digits(12),
                                                f"{cp2_first} {cp2_last}",
                                                self.rng.choice(BANK_KEYS), a.bank),
                                pid))
            self.tx(a, reactivate + timedelta(days=self.rng.randint(18, 35)),
                    "debit", self.amount(900, 4_700, 10),
                    self._narr_upi("debit", self.digits(12),
                                   rng.choice(MERCHANTS_UPI)[0] if (rng := self.rng) else "Merchant",
                                   "ICICI", "merchant@icici", a.bank))
            notes = "Long dormancy followed by a material reactivation burst."

        elif pid == 19:
            txs.extend(self.transfer(accounts[0], accounts[1], d,
                                     self.amount(260_000, 680_000, 1000), pid, "RTGS"))
            notes = "No prior relationship; first-ever contact is a large RTGS transfer."

        elif pid in (22, 23):
            for ai, acct_i in enumerate(accounts):
                unusual_day = d + timedelta(days=ai * 3)
                for i in range(self.rng.randint(7, 10)):
                    gap = self.rng.choice((1, 2, 9, 17, 31, 43))
                    unusual_day = min(unusual_day + timedelta(days=gap),
                                     self.period_end - timedelta(days=2))
                    direction = "credit" if i % 3 == 0 else "debit"
                    cp_f, cp_l = self._pick_name()
                    cp_bank = self.rng.choice(BANK_KEYS)
                    narr = self._narr_upi(direction, self.digits(12),
                                         f"{cp_f} {cp_l}",
                                         BANKS[cp_bank].upi_short,
                                         self._upi_handle(cp_f, cp_l, BANKS[cp_bank]),
                                         acct_i.bank)
                    txs.append(self.tx(acct_i, unusual_day, direction,
                                       self.amount(613, 9_743, 7),
                                       narr, pid, odd_time=True))
            tier = "safety_net"
            notes = ("Expected zero strong/weak findings from Patterns 1-19; "
                     "surface only via Pattern 22/23 safety-net trigger.")
        else:
            raise ValueError(f"Unknown pattern id {pid}")

        if not notes:
            notes = (f"Randomised {PATTERN_NAMES[pid]} evidence embedded in "
                     "realistic surrounding activity.")

        involved = (accounts if pid in (7, 8, 10, 12, 13, 16, 17, 19, 22, 23)
                    else [a])
        values = [float(t.amount) for t in txs]
        return {
            "pattern_id": pid, "pattern_name": PATTERN_NAMES[pid],
            "expected_tier": tier,
            "accounts_involved": [x.sid for x in involved],
            "_txn_objects": txs,
            "expected_amount_range": [min(values), max(values)] if values else [],
            "notes": notes,
        }

    def account_count(self, pid: int) -> int:
        if pid == 7: return self.rng.randint(3, 5)
        if pid == 8: return self.rng.randint(2, 4)
        if pid == 12: return self.rng.randint(7, 9)
        if pid == 16: return self.rng.randint(2, 3)
        if pid == 17: return self.rng.randint(2, 3)
        if pid in (10, 13, 19): return 2
        if pid in (22, 23): return self.rng.randint(2, 3)
        return 1

    # ------------------------------------------------------------------
    # Finalise balances
    # ------------------------------------------------------------------
    def finalize(self, accounts: list[Account]) -> None:
        for acct in accounts:
            acct.transactions.sort(key=lambda t: (t.when, t.bank_ref, t.direction))
            bal = acct.opening_balance
            minimum = bal
            for i, tx in enumerate(acct.transactions, 1):
                bal += tx.amount if tx.direction == "credit" else -tx.amount
                minimum = min(minimum, bal)
                tx.balance = bal
                tx.synthetic_ref = f"{acct.sid}_{i:06d}"
            if minimum < 0:
                adj = -minimum + self.amount(15_000, 45_000, 100)
                acct.opening_balance += adj
                bal = acct.opening_balance
                for tx in acct.transactions:
                    bal += tx.amount if tx.direction == "credit" else -tx.amount
                    tx.balance = bal

    # ------------------------------------------------------------------
    # Write account files (ONE file per account)
    # ------------------------------------------------------------------
    def write_account(self, folder: Path, acct: Account) -> None:
        txs = acct.transactions
        from_d = min(t.when.date() for t in txs) if txs else self.period_start
        to_d = max(t.when.date() for t in txs) if txs else self.period_end
        # Realistic file naming per bank style
        bp = acct.bank
        if bp.code == "SBI":
            fname = f"statement-{acct.account_no}.{acct.fmt}"
        elif bp.code == "HDFC":
            fname = f"{acct.account_no} statement.{acct.fmt}"
        elif bp.code in ("FEDERAL", "BOI", "BOB", "UCO"):
            fname = (f"{acct.account_no}-"
                     f"{from_d.strftime('%d-%m-%Y')}to{to_d.strftime('%d-%m-%Y')}"
                     f".{acct.fmt}")
        elif bp.code == "BANDHAN":
            fname = f"{acct.account_no}_SOA.{acct.fmt}"
        else:
            fname = f"{acct.account_no}_statement.{acct.fmt}"

        path = folder / fname
        sorted_txs = sorted(txs, key=lambda t: t.when)
        getattr(self, f"write_{acct.fmt}")(path, acct, sorted_txs)
        acct.files.append(f"statements/{fname}")

    @staticmethod
    def _dr(tx: Transaction) -> str:
        return f"{tx.amount:,.2f}" if tx.direction == "debit" else ""

    @staticmethod
    def _cr(tx: Transaction) -> str:
        return f"{tx.amount:,.2f}" if tx.direction == "credit" else ""

    # ------------------------------------------------------------------
    # CSV writers
    # ------------------------------------------------------------------
    def write_csv(self, path: Path, acct: Account, txs: list[Transaction]) -> None:
        bp = acct.bank
        from_d = txs[0].when.strftime("%d-%m-%Y") if txs else ""
        to_d = txs[-1].when.strftime("%d-%m-%Y") if txs else ""
        with path.open("w", newline="", encoding="utf-8-sig") as fh:
            w = csv.writer(fh, lineterminator="\n")
            # Header block — account metadata before the transaction table
            w.writerow(["Bank Name", bp.name])
            w.writerow(["Account Number", acct.account_no])
            w.writerow(["Account Holder Name", acct.name])
            w.writerow(["Branch", acct.branch])
            w.writerow(["IFSC Code", acct.ifsc])
            w.writerow(["Statement Period", f"{from_d} To {to_d}"])
            w.writerow([])
            # Use format matching OBS_CSV_009 (SBI-like CSV)
            if bp.code == "SBI":
                w.writerow(["TRAN-DATE", "TRAN_PARTICULAR", "CHQ-NUM",
                             "WITHDRAWAL", "DEPOSIT", "BALANCE"])
                for tx in txs:
                    w.writerow([tx.when.strftime("%d-%m-%Y"), tx.narration,
                                 tx.bank_ref, self._dr(tx), self._cr(tx),
                                 f"{tx.balance:.2f}"])
            else:
                # Generic readable CSV
                w.writerow(["Date", "Description", "Ref/ChequeNo",
                             "Debit", "Credit", "Balance"])
                for tx in txs:
                    w.writerow([bp.fmt_date(tx.when.date()), tx.narration,
                                 tx.bank_ref, self._dr(tx), self._cr(tx),
                                 f"{tx.balance:.2f}"])

    # ------------------------------------------------------------------
    # TXT writer (SBI / PNB fixed-width pipe-delimited)
    # ------------------------------------------------------------------
    def write_txt(self, path: Path, acct: Account, txs: list[Transaction]) -> None:
        bp = acct.bank
        COLS = [
            ("Trans Dt",          10),
            ("Value Dt",          10),
            ("Transn ID",         16),
            ("Transaction Particulars", 50),
            ("Ins Number",        14),
            ("Debit",             14),
            ("Credit",            14),
            ("Balance",           14),
        ]
        sep = "+" + "+".join("-" * (w + 2) for _, w in COLS) + "+"

        def row(vals: list[str]) -> str:
            parts = []
            for (_, w), v in zip(COLS, vals):
                parts.append(f" {v[:w]:<{w}} ")
            return "|" + "|".join(parts) + "|"

        def hdr() -> str:
            return row([h for h, _ in COLS])

        with path.open("w", encoding="utf-8", newline="\r\n") as fh:
            # Bank header block
            fh.write(f"{bp.name} - Account Statement\r\n")
            fh.write(f"Account Number : {acct.account_no}\r\n")
            fh.write(f"Account Name   : {acct.name}\r\n")
            fh.write(f"Branch         : {acct.branch}\r\n")
            fh.write(f"IFSC Code      : {acct.ifsc}\r\n")
            from_d = txs[0].when.date() if txs else self.period_start
            to_d = txs[-1].when.date() if txs else self.period_end
            fh.write(f"Statement Period: {from_d.strftime('%d-%m-%Y')} To "
                     f"{to_d.strftime('%d-%m-%Y')}\r\n")
            fh.write("\r\n")
            fh.write(sep + "\r\n")
            fh.write(hdr() + "\r\n")
            fh.write(sep + "\r\n")
            for tx in txs:
                dr = f"{tx.amount:,.2f}" if tx.direction == "debit" else ""
                cr = f"{tx.amount:,.2f}" if tx.direction == "credit" else ""
                fh.write(row([
                    tx.when.strftime("%d-%m-%Y"),
                    tx.when.strftime("%d-%m-%Y"),
                    tx.bank_ref,
                    tx.narration,
                    tx.bank_ref[:14],
                    dr,
                    cr,
                    f"{tx.balance:,.2f}",
                ]) + "\r\n")
            fh.write(sep + "\r\n")

    # ------------------------------------------------------------------
    # XLS/XLSX writers
    # ------------------------------------------------------------------
    def write_xls(self, path: Path, acct: Account, txs: list[Transaction]) -> None:
        """xlwt not available; emit .xlsx with .xlsx extension."""
        xlsx_path = Path(str(path).replace(".xls", ".xlsx"))
        self.write_xlsx(xlsx_path, acct, txs)
        if acct.files:
            acct.files[-1] = acct.files[-1].replace(".xls", ".xlsx")

    def write_xlsx(self, path: Path, acct: Account, txs: list[Transaction]) -> None:
        bp = acct.bank
        wb = Workbook()
        ws = wb.active
        ws.title = "Account Statement"

        hdr_fill = PatternFill("solid", fgColor="1F4E78")
        hdr_font = Font(bold=True, color="FFFFFF")

        if bp.xlsx_col_style == "federal_xls":
            cols = ["ACCOUNT NO.", "TRAN DATE", "VALUE DATE", "TRAN PARTICULAR",
                    "INSTRUMENT NO", "DEBIT AMOUNT", "CREDIT AMOUNT",
                    "BALANCE AMOUNT", "BALANCE INDICATOR", "ACCOUNT NAME",
                    "SOL ID", "SOL_DESC", "TRAN ID", "TRAN AMT",
                    "TRAN TYPE", "TRAN SUB TYPE"]
            ws.append(cols)
            for cell in ws[1]:
                cell.font = hdr_font; cell.fill = hdr_fill
            for tx in txs:
                sol_id = acct.ifsc[5:]
                ws.append([
                    acct.account_no,
                    tx.when.strftime("%d-%m-%Y"),
                    tx.when.strftime("%d-%m-%Y"),
                    tx.narration, tx.bank_ref,
                    float(tx.amount) if tx.direction == "debit" else None,
                    float(tx.amount) if tx.direction == "credit" else None,
                    float(tx.balance),
                    "C" if tx.direction == "credit" else "D",
                    acct.name, sol_id, acct.branch,
                    tx.bank_ref, float(tx.amount), "T",
                    "CI" if tx.direction == "credit" else "BI",
                ])
        elif bp.xlsx_col_style == "sbi_xls":
            cols = ["Ac_No", "AC_Name", "Tran_ID", "Tran_Date",
                    "Tran_Type", "Sub_Type", "Inst_Type", "Inst_Num",
                    "Dr_Amt", "Cr_Amt", "Balance", "Rmks", "Narration",
                    "pstd_dt", "Crncy", "value_dt"]
            ws.append(cols)
            for cell in ws[1]:
                cell.font = hdr_font; cell.fill = hdr_fill
            for tx in txs:
                ws.append([
                    acct.account_no, acct.name,
                    tx.bank_ref, tx.when.strftime("%d-%m-%Y"), "T",
                    "CI" if tx.direction == "credit" else "BI",
                    "E", tx.bank_ref,
                    float(tx.amount) if tx.direction == "debit" else None,
                    float(tx.amount) if tx.direction == "credit" else None,
                    float(tx.balance), tx.narration, tx.narration,
                    tx.when.strftime("%d-%m-%Y"), "INR",
                    tx.when.strftime("%d-%m-%Y"),
                ])
        else:
            # Generic XLSX matching OBS_XLSX_039 style
            # Header block rows — account metadata so parsers can identify the account
            from_d = txs[0].when.strftime("%d-%m-%Y") if txs else ""
            to_d   = txs[-1].when.strftime("%d-%m-%Y") if txs else ""
            ws.append(["Bank Name", bp.name])
            ws.append(["Account Number", acct.account_no])
            ws.append(["Account Holder Name", acct.name])
            ws.append(["Branch", acct.branch])
            ws.append(["IFSC Code", acct.ifsc])
            ws.append(["Statement Period", f"{from_d} To {to_d}"])
            ws.append([])
            # CTR BATCH NO | TXN DT | POST DATE | TXN TYPE | REF CHQ NO |
            # NARRATION | REF TXN NO | TXN BRANCH | DEBIT | CREDIT | BALANCE
            cols = ["CTR BATCH NO", "TXN DT", "POST DATE", "TXN TYPE",
                    "REF CHQ NO", "NARRATION", "REF TXN NO", "TXN BRANCH",
                    "DEBIT", "CREDIT", "BALANCE"]
            ws.append(cols)
            col_header_row = ws.max_row
            for cell in ws[col_header_row]:
                cell.font = hdr_font; cell.fill = hdr_fill
            batch = 1
            for tx in txs:
                ws.append([
                    batch, tx.when.strftime("%d-%m-%Y"),
                    tx.when.strftime("%d-%m-%Y"), "T", tx.bank_ref,
                    tx.narration, tx.bank_ref, acct.branch,
                    float(tx.amount) if tx.direction == "debit" else None,
                    float(tx.amount) if tx.direction == "credit" else None,
                    float(tx.balance),
                ])
                batch += 1

        ws.freeze_panes = "A2"
        for col, w in (("A", 18), ("B", 16), ("D", 50), ("F", 50)):
            try:
                ws.column_dimensions[col].width = w
            except Exception:
                pass
        wb.save(path)

    # ------------------------------------------------------------------
    # PDF writers
    # ------------------------------------------------------------------
    def write_pdf(self, path: Path, acct: Account, txs: list[Transaction]) -> None:
        style = acct.bank.pdf_col_style
        if style == "hdfc":
            self._write_pdf_hdfc(path, acct, txs)
        elif style == "axis":
            self._write_pdf_axis(path, acct, txs)
        elif style in ("sbi", "pnb", "boi", "bob", "uco"):
            self._write_pdf_sbi(path, acct, txs)
        elif style == "bandhan":
            self._write_pdf_bandhan(path, acct, txs)
        else:
            # kotak / federal / default
            self._write_pdf_kotak(path, acct, txs)

    def _common_doc(self, path: Path) -> SimpleDocTemplate:
        return SimpleDocTemplate(
            str(path), pagesize=landscape(A4),
            leftMargin=10*mm, rightMargin=10*mm,
            topMargin=10*mm, bottomMargin=16*mm,
        )

    def _footer_fn(self, bank_name: str, ifsc: str):
        def _footer(canvas, doc):
            canvas.saveState()
            canvas.setFont("Helvetica", 7)
            canvas.drawString(10*mm, 8*mm,
                f"This is a computer generated statement. {bank_name}  IFSC: {ifsc}")
            canvas.drawRightString(
                landscape(A4)[0] - 10*mm, 8*mm,
                f"Page {canvas.getPageNumber()}")
            canvas.restoreState()
        return _footer

    def _write_pdf_sbi(self, path: Path, acct: Account, txs: list[Transaction]) -> None:
        styles = getSampleStyleSheet()
        small = ParagraphStyle("small", parent=styles["Normal"], fontSize=6.5)
        doc = self._common_doc(path)
        story = []

        # Header
        bp = acct.bank
        period_from = min(t.when for t in txs).strftime("%d-%m-%Y") if txs else ""
        period_to = max(t.when for t in txs).strftime("%d-%m-%Y") if txs else ""
        hdr = [
            [f"{bp.name}", "", "", ""],
            [f"{acct.branch}", "", "", ""],
            ["", "", "", ""],
            [f"Account No : {acct.account_no}", f"CIF No : {self.digits(11)}",
             f"IFSC Code : {acct.ifsc}", f"MICR Code : {self.digits(9)}"],
            [f"Statement From : {period_from}", f"To : {period_to}",
             f"Account Status : OPEN", "Currency : INR"],
            [f"Opening Balance : {acct.opening_balance:,.2f} CR", "",
             f"Cleared Balance : {txs[-1].balance:,.2f} CR" if txs else "", ""],
        ]
        ht = Table(hdr, colWidths=[75*mm, 75*mm, 65*mm, 55*mm])
        ht.setStyle(TableStyle([
            ("SPAN", (0, 0), (3, 0)), ("SPAN", (0, 1), (3, 1)), ("SPAN", (0, 2), (3, 2)),
            ("FONTNAME", (0, 0), (3, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("GRID", (0, 3), (-1, -1), .3, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story += [ht, Spacer(1, 3*mm)]

        # Transactions table
        # Columns: Post Date | Value Date | Description | Cheque No | Debit | Credit | Balance
        rows = [["Post Date", "Value Date", "Description", "Cheque No/Ref", "Debit", "Credit", "Balance"]]
        for tx in txs:
            rows.append([
                tx.when.strftime("%d-%m-%Y"),
                tx.when.strftime("%d-%m-%Y"),
                Paragraph(tx.narration.replace("\n", "<br/>"), small),
                tx.bank_ref,
                self._dr(tx), self._cr(tx),
                f"{tx.balance:,.2f}",
            ])
        tbl = Table(rows, repeatRows=1,
                    colWidths=[22*mm, 22*mm, 100*mm, 30*mm, 24*mm, 24*mm, 28*mm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#003087")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 6.5),
            ("GRID", (0, 0), (-1, -1), .25, colors.lightgrey),
            ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(tbl)
        doc.build(story,
                  onFirstPage=self._footer_fn(bp.name, acct.ifsc),
                  onLaterPages=self._footer_fn(bp.name, acct.ifsc))

    def _write_pdf_hdfc(self, path: Path, acct: Account, txs: list[Transaction]) -> None:
        styles = getSampleStyleSheet()
        small = ParagraphStyle("small", parent=styles["Normal"], fontSize=6.5)
        doc = self._common_doc(path)
        story = []

        bp = acct.bank
        period_from = min(t.when for t in txs).strftime("%d/%m/%y") if txs else ""
        period_to = max(t.when for t in txs).strftime("%d/%m/%y") if txs else ""

        hdr = [
            [f"HDFC BANK LTD", "", "", ""],
            [f"Account Branch : {acct.branch}", "", "", ""],
            [f"MR/MS {acct.name}", f"City : {acct.city.upper()}", "", ""],
            [f"RTGS/NEFT IFSC : {acct.ifsc}", f"MICR: {self.digits(9)}", "", ""],
            [f"Cust ID : {self.digits(9)}", f"Account No : {acct.account_no}",
             f"A/C Open Date : {(self.period_start - timedelta(days=30)).strftime('%d/%m/%Y')}",
             "Currency : INR"],
            [f"Statement From: {period_from}", f"To: {period_to}",
             f"Account Status : Active", ""],
        ]
        ht = Table(hdr, colWidths=[75*mm, 75*mm, 65*mm, 55*mm])
        ht.setStyle(TableStyle([
            ("SPAN", (0, 0), (3, 0)), ("SPAN", (0, 1), (3, 1)),
            ("FONTNAME", (0, 0), (3, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("GRID", (0, 4), (-1, -1), .3, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story += [ht, Spacer(1, 3*mm)]

        # Columns: Date | Narration | Chq./Ref.No. | Value Dt | Withdrawal | Deposit | Closing Balance
        rows = [["Date", "Narration", "Chq./Ref.No.", "Value Dt",
                 "Withdrawal Amt.", "Deposit Amt.", "Closing Balance"]]
        for tx in txs:
            rows.append([
                tx.when.strftime("%d/%m/%y"),
                Paragraph(tx.narration.replace("\n", "<br/>"), small),
                tx.bank_ref,
                tx.when.strftime("%d/%m/%y"),
                self._dr(tx), self._cr(tx),
                f"{tx.balance:,.2f}",
            ])
        tbl = Table(rows, repeatRows=1,
                    colWidths=[18*mm, 110*mm, 28*mm, 18*mm, 24*mm, 24*mm, 28*mm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#004C97")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 6.5),
            ("GRID", (0, 0), (-1, -1), .25, colors.lightgrey),
            ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(tbl)
        doc.build(story,
                  onFirstPage=self._footer_fn(bp.name, acct.ifsc),
                  onLaterPages=self._footer_fn(bp.name, acct.ifsc))

    def _write_pdf_axis(self, path: Path, acct: Account, txs: list[Transaction]) -> None:
        styles = getSampleStyleSheet()
        small = ParagraphStyle("small", parent=styles["Normal"], fontSize=6.5)
        doc = self._common_doc(path)
        story = []

        bp = acct.bank
        period_from = min(t.when for t in txs).strftime("%d-%m-%Y") if txs else ""
        period_to = max(t.when for t in txs).strftime("%d-%m-%Y") if txs else ""

        hdr = [
            ["AXIS BANK LIMITED", "", "", ""],
            [f"Customer No : {self.digits(9)}", f"Scheme : CA - BUSINESS ADVANTAGE", "", ""],
            [f"Statement of Account No : {acct.account_no}", f"for the period (From : {period_from} To : {period_to})", "", ""],
            [f"Branch : {acct.branch}", f"Currency : INR", f"IFSC : {acct.ifsc}", ""],
        ]
        ht = Table(hdr, colWidths=[75*mm, 90*mm, 55*mm, 50*mm])
        ht.setStyle(TableStyle([
            ("SPAN", (0, 0), (3, 0)),
            ("FONTNAME", (0, 0), (3, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("GRID", (0, 3), (-1, 3), .3, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story += [ht, Spacer(1, 3*mm)]

        # Columns: Tran Date | Chq No | Particulars | Debit | Credit | Balance | Init. | Br
        rows = [["Tran Date", "Chq No", "Particulars", "Debit", "Credit", "Balance", "Init.", "Br"]]
        for tx in txs:
            rows.append([
                tx.when.strftime("%d-%m-%Y"),
                tx.bank_ref[:8],
                Paragraph(tx.narration.replace("\n", "<br/>"), small),
                self._dr(tx), self._cr(tx),
                f"{tx.balance:,.2f}",
                self.alpha(3),
                self.digits(4),
            ])
        tbl = Table(rows, repeatRows=1,
                    colWidths=[22*mm, 24*mm, 100*mm, 22*mm, 22*mm, 24*mm, 12*mm, 10*mm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#800000")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 6.5),
            ("GRID", (0, 0), (-1, -1), .25, colors.lightgrey),
            ("ALIGN", (3, 1), (5, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(tbl)
        doc.build(story,
                  onFirstPage=self._footer_fn(bp.name, acct.ifsc),
                  onLaterPages=self._footer_fn(bp.name, acct.ifsc))

    def _write_pdf_kotak(self, path: Path, acct: Account, txs: list[Transaction]) -> None:
        styles = getSampleStyleSheet()
        small = ParagraphStyle("small", parent=styles["Normal"], fontSize=6.5)
        doc = self._common_doc(path)
        story = []

        bp = acct.bank
        period_from = min(t.when for t in txs).strftime("%d-%m-%Y") if txs else ""
        period_to = max(t.when for t in txs).strftime("%d-%m-%Y") if txs else ""

        hdr_data = [
            [f"{bp.name}", "", "", ""],
            [f"Account No : {acct.account_no}",
             f"Cust.Reln.No : {self.digits(9)}",
             f"IFSC Code : {acct.ifsc}",
             f"MICR Code : {self.digits(9)}"],
            [f"Branch : {acct.branch}", f"City : {acct.city.upper()}", "", ""],
            [f"Period : {period_from} To {period_to}",
             f"Currency : INR",
             f"Opening Balance : {acct.opening_balance:,.2f}(Cr)", ""],
        ]
        ht = Table(hdr_data, colWidths=[75*mm, 75*mm, 65*mm, 55*mm])
        ht.setStyle(TableStyle([
            ("SPAN", (0, 0), (3, 0)),
            ("FONTNAME", (0, 0), (3, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("BOX", (0, 0), (-1, -1), .5, colors.grey),
            ("INNERGRID", (0, 1), (-1, -1), .25, colors.lightgrey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story += [ht, Spacer(1, 3*mm)]

        # Columns: Date | Narration | Chq/Ref No | Withdrawal (Dr) | Deposit(Cr) | Balance
        rows = [["Date", "Narration", "Chq/Ref No",
                 "Withdrawal (Dr)", "Deposit(Cr)", "Balance"]]
        for tx in txs:
            rows.append([
                tx.when.strftime("%d-%m-%Y"),
                Paragraph(tx.narration.replace("\n", "<br/>"), small),
                tx.bank_ref,
                self._dr(tx), self._cr(tx),
                f"{tx.balance:,.2f}",
            ])
        tbl = Table(rows, repeatRows=1,
                    colWidths=[22*mm, 115*mm, 30*mm, 26*mm, 26*mm, 28*mm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#CC0000")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 6.5),
            ("GRID", (0, 0), (-1, -1), .25, colors.lightgrey),
            ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(tbl)
        doc.build(story,
                  onFirstPage=self._footer_fn(bp.name, acct.ifsc),
                  onLaterPages=self._footer_fn(bp.name, acct.ifsc))

    def _write_pdf_bandhan(self, path: Path, acct: Account, txs: list[Transaction]) -> None:
        styles = getSampleStyleSheet()
        small = ParagraphStyle("small", parent=styles["Normal"], fontSize=6.5)
        doc = self._common_doc(path)
        story = []

        bp = acct.bank
        period_from = min(t.when for t in txs).strftime("%d-%b-%Y") if txs else ""
        period_to = max(t.when for t in txs).strftime("%d-%b-%Y") if txs else ""

        hdr = [
            ["BANDHAN BANK LIMITED", "", "", ""],
            [f"Customer Number: {self.digits(9)}", f"Branch of Ownership: {acct.branch}", "", ""],
            [f"Account No: {acct.account_no}", f"IFSC: {acct.ifsc}", "", ""],
            [f"Account Title: {acct.name}",
             f"Branch ID: {self.digits(4)}",
             f"MICR: {self.digits(9)}", ""],
            [f"Period: {period_from} to {period_to}", "Currency: INR", "", ""],
        ]
        ht = Table(hdr, colWidths=[75*mm, 75*mm, 65*mm, 55*mm])
        ht.setStyle(TableStyle([
            ("SPAN", (0, 0), (3, 0)),
            ("FONTNAME", (0, 0), (3, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("BOX", (0, 0), (-1, -1), .5, colors.grey),
            ("INNERGRID", (0, 1), (-1, -1), .25, colors.lightgrey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story += [ht, Spacer(1, 3*mm)]

        # Kotak-like column style for Bandhan
        rows = [["Date", "Narration", "Chq/Ref No",
                 "Withdrawal (Dr)", "Deposit(Cr)", "Balance"]]
        for tx in txs:
            rows.append([
                tx.when.strftime("%d-%b-%Y"),
                Paragraph(tx.narration.replace("\n", "<br/>"), small),
                tx.bank_ref,
                self._dr(tx), self._cr(tx),
                f"{tx.balance:,.2f}",
            ])
        tbl = Table(rows, repeatRows=1,
                    colWidths=[24*mm, 113*mm, 30*mm, 26*mm, 26*mm, 28*mm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E07B39")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 6.5),
            ("GRID", (0, 0), (-1, -1), .25, colors.lightgrey),
            ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(tbl)
        doc.build(story,
                  onFirstPage=self._footer_fn(bp.name, acct.ifsc),
                  onLaterPages=self._footer_fn(bp.name, acct.ifsc))

    # ------------------------------------------------------------------
    # Serialise ground truth
    # ------------------------------------------------------------------
    def _serialise_finding(self, f: dict[str, Any]) -> dict[str, Any]:
        txns = f.pop("_txn_objects")
        f["expected_txn_refs"] = [t.bank_ref for t in txns]
        f["synthetic_txn_refs"] = [t.synthetic_ref for t in txns]
        return f

    def write_case(self, root: Path, folder_name: str, accounts: list[Account],
                   findings: list[dict[str, Any]], description: str,
                   expected_non_findings: list[dict[str, Any]] | None = None,
                   extra: dict[str, Any] | None = None) -> None:
        folder = root / folder_name
        stmts = folder / "statements"
        stmts.mkdir(parents=True, exist_ok=True)
        self.finalize(accounts)
        for acct in accounts:
            self.write_account(stmts, acct)
        truth: dict[str, Any] = {
            "schema_version": 1, "folder": folder_name,
            "description": description, "seed": self.seed,
            "accounts": [
                {"synthetic_account_id": a.sid, "role": a.role,
                 "file": a.files[0] if a.files else "", "files": a.files,
                 "fabricated_account_number": a.account_no,
                 "bank": a.bank.name}
                for a in accounts
            ],
            "expected_findings": [self._serialise_finding(x) for x in findings],
            "expected_non_findings": expected_non_findings or [],
        }
        if extra:
            truth.update(extra)
        (folder / "ground_truth.json").write_text(
            json.dumps(truth, indent=2) + "\n", encoding="utf-8")

    # ------------------------------------------------------------------
    # Generate clean control
    # ------------------------------------------------------------------
    def generate_clean(self, root: Path) -> None:
        accounts = [self.new_account("clean_control")
                    for _ in range(self.rng.randint(4, 6))]
        for acct in accounts:
            d = self.event_day()
            sal = self.amount(48_000, 112_000, 500)
            self.tx(acct, d, "credit", sal,
                    self._narr_salary(self.rng.choice(EMPLOYERS),
                                      self.ref("S", 14), acct.bank))
            self.tx(acct, d + timedelta(days=1), "debit",
                    self.amount(2_000, 8_000, 500, False),
                    self._narr_atm(acct))
            if self.rng.random() < .65:
                ref = self.ref("U", 12)
                failed = self.tx(acct, d + timedelta(days=8), "debit",
                                 self.amount(700, 3_500, 10),
                                 self._narr_upi("debit", self.digits(12),
                                                "Bill Payment", "PAYTM",
                                                "billpay@paytm", acct.bank))
                self.tx(acct, d + timedelta(days=9), "credit", failed.amount,
                        f"RETURN/{failed.bank_ref}/REVERSAL")
        extra = {
            "tier_expectations": {
                "strong_findings": 0,
                "weak_findings_allowed": True,
                "ranking_expectation": "clean accounts must remain outside the suspicious top ranks",
            }
        }
        non = [{"pattern_id": p,
                "reason": "No deliberately planted strong-tier behaviour."}
               for p in PATTERN_NAMES if p < 20]
        self.write_case(root, "clean_control", accounts, [],
                        "Negative control with normal banking friction and weak-tier noise.",
                        non, extra)

    # ------------------------------------------------------------------
    # Generate per-pattern standalone folders
    # ------------------------------------------------------------------
    def generate_standalone(self, root: Path) -> None:
        for pid in PATTERN_NAMES:
            count = self.account_count(pid)
            roles = ["subject"] * count
            if pid == 12:
                roles = ["hub"] + ["spoke"] * (count - 1)
            elif pid in (7, 8, 17):
                roles = ["originator"] + ["intermediary"] * (count - 2) + ["recipient"]
            elif pid in (10, 13, 19):
                roles = ["originator", "counterparty"]
            primary = [self.new_account(roles[i], normal=(pid != 18))
                       for i in range(count)]
            finding = self.plant(pid, primary)
            noise_n = 0 if pid in (22, 23) else self.rng.randint(2, 5)
            noise = [self.new_account("clean_control") for _ in range(noise_n)]
            accounts = primary + noise
            non_pid = 7 if pid != 7 else 5
            non = ([] if pid in (22, 23)
                   else [{"pattern_id": non_pid,
                          "reason": "No unrelated fixture planted for this pattern."}])
            extra = None
            if pid in (22, 23):
                extra = {"safety_net_expectation": (
                    "Expected zero strong/weak findings from Patterns 1-19; "
                    "surface only via Pattern 22/23 safety-net trigger."
                )}
                non = [{"pattern_id": x,
                        "reason": "Unknown-shape fixture avoids written rule thresholds."}
                       for x in PATTERN_NAMES if x < 20]
            self.write_case(root, FOLDER_NAMES[pid], accounts, [finding],
                            finding["notes"], non, extra)

    # ------------------------------------------------------------------
    # Generate combined all-patterns folder
    # ------------------------------------------------------------------
    def generate_combined(self, root: Path) -> None:
        accounts = [self.new_account("mixed_subject") for _ in range(24)]
        accounts.extend(self.new_account("clean_control") for _ in range(72))
        groups = {
            1: [0], 2: [1], 3: [2], 4: [3], 5: [4], 7: [5, 6, 7],
            8: [8, 9, 10], 9: [11], 10: [12, 13], 11: [14],
            12: [5, 6, 7, 8, 9, 10, 11], 13: [15, 16], 14: [17],
            15: [18], 16: [12, 13, 14], 17: [0, 1, 2], 18: [19],
            19: [3, 4], 22: [20], 23: [21],
        }
        findings = [self.plant(pid, [accounts[i] for i in idx])
                    for pid, idx in groups.items()]
        for i in (22, 23):
            accounts[i].role = "clean_control"
        weights: dict[str, int] = {a.sid: 0 for a in accounts}
        for f in findings:
            sc = 3 if f["expected_tier"] == "strong" else 1
            for sid in f["accounts_involved"]:
                weights[sid] += sc
        ranking = [sid for sid, _ in sorted(weights.items(),
                                             key=lambda x: (-x[1], x[0]))]
        extra = {
            "expected_account_ranking": ranking,
            "ranking_basis": "Descending planted-evidence weight; strong=3, weak/safety-net=1.",
            "pure_clean_control_accounts": [a.sid for a in accounts[24:]],
            "noise_population_count": len(accounts) - 24,
            "pattern_21_validation": (
                "Compare pipeline suspicious-account ranking with "
                "expected_account_ranking; Pattern 21 is composite and not planted."
            ),
        }
        self.write_case(
            root, "combined_all_patterns", accounts, findings,
            "All rule fixtures and both unknown-shape safety nets woven through "
            "96 accounts including clean controls.", [], extra)

    # ------------------------------------------------------------------
    # Sample mode: one account per bank
    # ------------------------------------------------------------------
    def generate_samples(self, root: Path) -> None:
        sample_dir = root / "_bank_samples"
        if sample_dir.exists():
            shutil.rmtree(sample_dir)
        sample_dir.mkdir(parents=True)
        for bank_code, bp in BANKS.items():
            # Create one account with ~4 months of activity
            orig_end = self.period_end
            orig_start = self.period_start
            self.period_end = self.today - timedelta(days=30)
            self.period_start = self.period_end - timedelta(days=120)
            acct = self.new_account("subject", normal=True, bank_code=bank_code)
            # Force to pdf for visual inspection
            acct.fmt = "pdf"
            self.period_end = orig_end
            self.period_start = orig_start
            self.finalize([acct])
            self.write_account(sample_dir, acct)
            print(f"  [{bank_code}] {bp.name}: {len(acct.transactions)} txns -> {acct.files[-1]}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate PII-free bank statement regression fixtures.")
    parser.add_argument("--output", type=Path,
                        default=Path(__file__).resolve().parent)
    parser.add_argument("--seed", type=int, default=None,
                        help="Reproducible seed (omit for random)")
    parser.add_argument("--keep-existing", action="store_true")
    parser.add_argument("--sample-only", action="store_true",
                        help="Generate one sample PDF per bank for review, then stop")
    args = parser.parse_args()

    seed = (args.seed if args.seed is not None
            else random.SystemRandom().randrange(1, 2**63))
    root = args.output.resolve()
    root.mkdir(parents=True, exist_ok=True)

    gen = Generator(seed)

    if args.sample_only:
        print(f"Generating bank samples (seed={seed}) ...")
        gen.generate_samples(root)
        print(f"\nSamples written to: {root / '_bank_samples'}")
        print("Review the PDFs, then re-run without --sample-only to generate the full dataset.")
        return 0

    generated = ["clean_control", "combined_all_patterns", *FOLDER_NAMES.values()]
    if not args.keep_existing:
        for name in generated:
            p = root / name
            if p.exists():
                shutil.rmtree(p)

    print(f"Generating full dataset (seed={seed}) ...")
    gen.generate_clean(root)
    gen.generate_standalone(root)
    gen.generate_combined(root)

    manifest = {
        "schema_version": 1, "seed": seed,
        "generated_on": date.today().isoformat(),
        "folders": ["clean_control", *FOLDER_NAMES.values(), "combined_all_patterns"],
    }
    (root / "generation_manifest.json").write_text(
        json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    print(f"Generated {len(manifest['folders'])} fixture folders with seed {seed}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
